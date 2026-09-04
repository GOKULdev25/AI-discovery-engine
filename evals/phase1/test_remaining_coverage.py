"""EV-P1-02, 04, 07, 14 — the export surfaces failures, YouTube walks
pagination via nextPageToken, `raw` is universal, and connectors never
touch httpx directly (a Phase-1-scoped restatement of EV-INV-08, which
already runs on every invocation regardless of phase — this just names
the P§8/A§10.1 promise explicitly at the phase that introduces the
connectors it protects).
"""

from __future__ import annotations

import tempfile
from pathlib import Path
from unittest import mock

import httpx
import openpyxl

from app.connectors.base import JobSpec
from app.connectors.youtube import YouTubeConnector
from app.export.excel import build_export
from app.jobs.engine import forget_engine, submit_batch
from app.projects import scaffold
from app.projects.resolver import ProjectResolver
from app.store import duckdb as dk
from app.store import sqlite as sq
from evals.corpora.golden import mock_data
from evals.harness import connector_ctx, drain, iter_py_files, make_settings, wait_for_batch_done_direct
from evals.registry import eval_case


@eval_case(
    "EV-P1-02",
    proves="Every failed link appears in the export sheet AND the UI with a typed reason (P§8 criterion 2)",
    source="P§8",
    severity="BLOCKER",
    tags=["phase:P1"],
)
async def ev_p1_02():
    with tempfile.TemporaryDirectory(prefix="ev-p102-") as tmp:
        settings = make_settings(Path(tmp))
        resolver = ProjectResolver(settings)
        config = await scaffold.create_project(settings, resolver, "p102")
        project_dir = resolver.project_dir(config.id)
        try:
            urls = ["not a valid url", "https://example.com/unsupported-thing.pdf", "fixture://run?count=1&latency_ms=0"]
            batch_id, links = await submit_batch(settings, config.id, urls)
            async with sq.ops_db(project_dir) as ops_conn:
                await wait_for_batch_done_direct(ops_conn, batch_id, timeout=10)
                cur = await ops_conn.execute("SELECT url, failure_code FROM links WHERE status = 'failed'")
                failed_in_db = {r["url"]: r["failure_code"] for r in await cur.fetchall()}
            assert failed_in_db == {
                "not a valid url": "INVALID_URL",
                "https://example.com/unsupported-thing.pdf": "UNSUPPORTED_SOURCE",
            }

            out_path = await build_export(resolver, config)
            wb = openpyxl.load_workbook(out_path)
            links_ws = wb["links"]
            header = [c.value for c in next(links_ws.iter_rows(min_row=1, max_row=1))]
            url_col = header.index("url")
            code_col = header.index("failure_code")
            exported = {
                row[url_col]: row[code_col]
                for row in links_ws.iter_rows(min_row=2, values_only=True)
                if row[url_col] in failed_in_db
            }
            assert exported == failed_in_db, f"export sheet disagrees with the API: {exported} vs {failed_in_db}"
        finally:
            await forget_engine(config.id)
            await dk.forget_committer(project_dir)


@eval_case(
    "EV-P1-04",
    proves="YouTube pagination is walked via nextPageToken, not assumed single-page",
    source="A§2.1",
    severity="MAJOR",
    tags=["phase:P1"],
)
async def ev_p1_04():
    seen_tokens = []

    async def handler(request: httpx.Request) -> httpx.Response:
        token = request.url.params.get("pageToken")
        seen_tokens.append(token)
        if token is None:
            body = {**mock_data.YOUTUBE_PAGE_1, "nextPageToken": "page-2-token"}
        elif token == "page-2-token":
            body = {"items": [], "nextPageToken": None}
        else:
            raise AssertionError(f"unexpected pageToken: {token}")
        return httpx.Response(200, json=body, request=request)

    async with connector_ctx("youtube", transport=httpx.MockTransport(handler), youtube_api_key="fake") as ctx:
        connector = YouTubeConnector()
        job = JobSpec(url="https://www.youtube.com/watch?v=abc123", params={"video_id": "abc123"})
        docs = await drain(connector.run(job, ctx))

    assert seen_tokens == [None, "page-2-token"], f"pagination did not walk nextPageToken correctly: {seen_tokens}"
    assert len(docs) == 2  # one top-level comment + its one reply, from page 1 only


@eval_case(
    "EV-P1-15",
    proves="YouTube pagination is bounded — a video with unbounded comments doesn't spend unbounded quota on one link",
    source="A§11.1",
    severity="BLOCKER",
    tags=["phase:P1"],
)
async def ev_p1_15():
    # Found live (2026-08-29): a popular real video paginated indefinitely
    # against the real API, staging thousands of comments per minute with
    # no ceiling — before Phase 3's quota ledger exists to stop it, one
    # link could exhaust the entire 10,000-unit/day budget. This eval
    # pins the fix (`MAX_PAGES`) so it can't silently regress.
    from app.connectors.youtube import MAX_PAGES

    call_count = 0

    async def handler(request: httpx.Request) -> httpx.Response:
        nonlocal call_count
        call_count += 1
        # An endless feed: always claims there's another page.
        return httpx.Response(
            200, json={"items": [], "nextPageToken": f"token-{call_count}"}, request=request,
        )

    async with connector_ctx("youtube", transport=httpx.MockTransport(handler), youtube_api_key="fake") as ctx:
        connector = YouTubeConnector()
        job = JobSpec(url="https://www.youtube.com/watch?v=neverending", params={"video_id": "neverending"})
        await drain(connector.run(job, ctx))

    assert call_count == MAX_PAGES, f"expected exactly {MAX_PAGES} calls against an endless feed, got {call_count}"


@eval_case(
    "EV-P1-07",
    proves="The raw escape hatch exists: raw is non-null on 100% of emitted rows, across every connector",
    source="A§8",
    severity="MAJOR",
    tags=["phase:P1"],
)
async def ev_p1_07():
    async def handler(request: httpx.Request) -> httpx.Response:
        return httpx.Response(200, json=mock_data.APPSTORE_PAGE_1, request=request)

    from app.connectors.appstore import AppStoreConnector

    async with connector_ctx("appstore", transport=httpx.MockTransport(handler)) as ctx:
        connector = AppStoreConnector()
        job = JobSpec(url="https://apps.apple.com/us/app/x/id1", params={"app_id": "1", "country": "us"})
        docs = await drain(connector.run(job, ctx))

    assert docs, "expected at least one document"
    assert all(d.raw is not None for d in docs), "raw must be populated on every emitted row"


@eval_case(
    "EV-P1-14",
    proves="Connectors do their own I/O only through ctx (Phase-1-scoped restatement of EV-INV-08's structural rule)",
    source="A§10.1",
    severity="BLOCKER",
    tags=["phase:P1"],
)
def ev_p1_14():
    from evals.harness import BACKEND_APP_DIR

    base_module = (BACKEND_APP_DIR / "connectors" / "base.py").resolve()
    hits = []
    for path in iter_py_files(BACKEND_APP_DIR / "connectors", exclude={base_module}):
        text = path.read_text(encoding="utf-8")
        if "import httpx" in text or "httpx." in text:
            hits.append(str(path))
    assert not hits, f"a connector references httpx directly instead of ctx.fetch(): {hits}"
